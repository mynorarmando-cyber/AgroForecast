"""
Lógica del modelo estadístico de rendimientos y curva de cosecha.
Compartida entre la app de Streamlit y el uso por línea de comandos.

Metodología (nivel analista de datos):
- Se excluye el primer y el último ciclo registrado de cada Finca-Lote:
  el primer ciclo suele arrastrar captura incompleta de los primeros años
  del sistema, y el último ciclo suele estar en curso (semanas de cosecha
  aún no registradas). Ambos distorsionan curva y rendimiento si se dejan.
- Tendencia: regresión robusta Theil-Sen (Rendimiento ~ Año de inicio de
  cosecha) — no asume normalidad y es resistente a atípicos, por lo que no
  requiere recortar (winsorizar) datos a mano.
- Estacionalidad: se remueve la tendencia de cada ciclo (residuo), y se
  suaviza por semana del año con un kernel circular (la semana 53 conecta
  con la 1) ponderado por cercanía de semana y por recencia del año
  (decaimiento exponencial) — así todos los años aportan, pero los
  recientes pesan más, sin depender de un corte arbitrario entre "antes"
  y "después".
- Recomendado = tendencia proyectada al año objetivo + estacionalidad de
  esa semana.
"""
import io
import numpy as np
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

FONT = 'Calibri'
H_FILL = PatternFill('solid', fgColor='1F4E28')
SUB_FILL = PatternFill('solid', fgColor='D9E8D6')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# Carga y limpieza
# ---------------------------------------------------------------------------
def cargar_tabla6(file_like):
    t6 = pd.read_excel(file_like, sheet_name='Hoja1', header=1, usecols="B:N", nrows=14084)
    return t6


def marcar_ciclos_incompletos(t6):
    """Marca el primer y último Ciclo de cada Finca-Lote como incompletos."""
    rango = t6.groupby(['Finca', 'Lote'])['Ciclo'].agg(ciclo_min='min', ciclo_max='max')
    t6 = t6.merge(rango, on=['Finca', 'Lote'])
    t6['ciclo_incompleto'] = (t6['Ciclo'] == t6['ciclo_min']) | (t6['Ciclo'] == t6['ciclo_max'])
    return t6.drop(columns=['ciclo_min', 'ciclo_max'])


def opciones_disponibles(t6):
    vegetales = sorted(t6['Referencia'].dropna().unique().tolist())
    fincas = sorted(t6['Finca'].dropna().unique().tolist())
    return vegetales, fincas


# ---------------------------------------------------------------------------
# Construcción de ciclos y cortes (a nivel Finca-Lote-Ciclo-Referencia)
# ---------------------------------------------------------------------------
def construir_ciclos(t6, referencias_vegetal, fincas=None, excluir_incompletos=True):
    df = t6[t6['Referencia'].isin(referencias_vegetal)].copy()
    if fincas:
        df = df[df['Finca'].isin(fincas)]
    if excluir_incompletos:
        df = df[~df['ciclo_incompleto']]
    if df.empty:
        raise ValueError("No hay datos para esa combinación de vegetal/finca.")

    grp = df.groupby(['Finca', 'Lote', 'Ciclo', 'Codigo'], as_index=False).agg(
        Kilos=('Kilos', 'sum'), Semana=('Semana', 'first'),
        Año=('Año', 'first'), Area=('Area', 'first'), CantidadV=('Cantidad V', 'first')
    )
    grp = grp.sort_values(['Finca', 'Lote', 'Ciclo', 'Codigo'])
    grp['corte'] = grp.groupby(['Finca', 'Lote', 'Ciclo']).cumcount() + 1
    grp['cycle_total'] = grp.groupby(['Finca', 'Lote', 'Ciclo'])['Kilos'].transform('sum')
    grp['pct'] = grp['Kilos'] / grp['cycle_total']

    cyc = grp[grp.corte == 1][['Finca', 'Lote', 'Ciclo', 'Semana', 'Año', 'Area', 'CantidadV']].rename(
        columns={'Semana': 'SemanaInicio', 'Año': 'AñoInicio'})
    cyc = cyc.merge(grp.groupby(['Finca', 'Lote', 'Ciclo'])['Kilos'].sum().rename('TotalKilos'),
                     on=['Finca', 'Lote', 'Ciclo'])
    cyc['Rendimiento'] = cyc['TotalKilos'] / (cyc['Area'] * cyc['CantidadV'])
    cortes = grp.merge(cyc[['Finca', 'Lote', 'Ciclo']], on=['Finca', 'Lote', 'Ciclo'])
    return cyc, cortes


# ---------------------------------------------------------------------------
# Curva de cosecha por corte — global y por año
# ---------------------------------------------------------------------------
def curva_por_anio(cortes, maxcorte):
    """% promedio por corte, para cada año de inicio de cosecha. Incluye N."""
    sub = cortes[cortes.corte <= maxcorte].copy()
    tabla = sub.pivot_table(index='corte', columns='Año', values='pct', aggfunc='mean')
    tabla = tabla.reindex(range(1, maxcorte + 1))
    tabla_norm = tabla.div(tabla.sum(axis=0), axis=1)
    n_ciclos = sub.groupby('Año')['Ciclo'].nunique().reindex(tabla.columns)
    return tabla_norm, n_ciclos


def curva_recomendada(cortes, cyc, maxcorte, decay=0.85, anio_referencia=None):
    """Curva ponderada por recencia (todos los años, decaimiento exponencial)."""
    if anio_referencia is None:
        anio_referencia = cyc['AñoInicio'].max()
    sub = cortes[cortes.corte <= maxcorte].copy()
    sub['peso'] = decay ** (anio_referencia - sub['Año'])
    agg = sub.groupby('corte').apply(
        lambda g: np.average(g['pct'], weights=g['peso'])
    ).reindex(range(1, maxcorte + 1)).fillna(0)
    return agg / agg.sum()


# ---------------------------------------------------------------------------
# Tendencia robusta + estacionalidad ponderada por recencia
# ---------------------------------------------------------------------------
def circular_week_distance(w1, w2, n=53):
    d = np.abs(w1 - w2)
    return np.minimum(d, n - d)


def tendencia_robusta(cyc):
    """Theil-Sen: robusto a atípicos, no requiere winsorizar a mano."""
    slope, intercept, low, high = stats.theilslopes(cyc['Rendimiento'], cyc['AñoInicio'])
    cyc = cyc.copy()
    cyc['trend_pred'] = intercept + slope * cyc['AñoInicio']
    cyc['resid'] = cyc['Rendimiento'] - cyc['trend_pred']
    return cyc, {'slope': slope, 'intercept': intercept, 'slope_low': low, 'slope_high': high}


def estacionalidad_kernel(cyc_con_resid, window=3, decay=0.85, anio_referencia=None):
    """Estacionalidad semanal vía kernel circular (distancia de semana)
    ponderado por recencia (decaimiento exponencial por año). Usa TODOS
    los años disponibles, sin cortes arbitrarios entre periodos."""
    if anio_referencia is None:
        anio_referencia = cyc_con_resid['AñoInicio'].max()
    recency_w = decay ** (anio_referencia - cyc_con_resid['AñoInicio'])
    weeks_sem = cyc_con_resid['SemanaInicio'].values
    resid = cyc_con_resid['resid'].values
    recency_w = recency_w.values
    out = {}
    for w in range(1, 54):
        dist = circular_week_distance(weeks_sem, w)
        mask = dist <= window
        if not mask.any():
            out[w] = np.nan
            continue
        kernel_w = 1 - dist[mask] / (window + 1)
        weight = kernel_w * recency_w[mask]
        if weight.sum() <= 0:
            out[w] = np.nan
            continue
        out[w] = np.average(resid[mask], weights=weight)
    s = pd.Series(out).interpolate(limit_direction='both')
    return s


def rendimiento_real_por_semana_y_anio(cyc):
    """Tabla real (sin suavizar): rendimiento promedio por semana y año."""
    tabla = cyc.pivot_table(index='SemanaInicio', columns='AñoInicio', values='Rendimiento', aggfunc='mean')
    tabla = tabla.reindex(range(1, 54))
    n = cyc.pivot_table(index='SemanaInicio', columns='AñoInicio', values='Rendimiento', aggfunc='count')
    n = n.reindex(range(1, 54))
    return tabla, n


def rendimiento_recomendado(cyc, anio_objetivo, decay=0.85, window=3):
    cyc_r, trend_stats = tendencia_robusta(cyc)
    seasonal = estacionalidad_kernel(cyc_r, window=window, decay=decay)
    trend_target = trend_stats['intercept'] + trend_stats['slope'] * anio_objetivo
    recomendado = (trend_target + seasonal).clip(lower=0)
    trend_stats['trend_target'] = trend_target
    trend_stats['n_ciclos'] = len(cyc)
    # prueba de significancia de la tendencia (correlación robusta de rangos)
    rho, p = stats.spearmanr(cyc['AñoInicio'], cyc['Rendimiento'])
    trend_stats['spearman_rho'] = rho
    trend_stats['p_value'] = p
    return recomendado, cyc_r, trend_stats


# ---------------------------------------------------------------------------
# Orquestador
# ---------------------------------------------------------------------------
def ejecutar_modelo(t6, nombre_vegetal, referencias_vegetal, fincas=None,
                     maxcorte=3, anio_objetivo=None, decay=0.85, window=3,
                     excluir_incompletos=True):
    t6m = marcar_ciclos_incompletos(t6)
    cyc, cortes = construir_ciclos(t6m, referencias_vegetal, fincas, excluir_incompletos)
    if anio_objetivo is None:
        anio_objetivo = int(cyc['AñoInicio'].max()) + 1

    curva_tabla, curva_n = curva_por_anio(cortes, maxcorte)
    curva_rec = curva_recomendada(cortes, cyc, maxcorte, decay=decay, anio_referencia=cyc['AñoInicio'].max())

    rend_real_tabla, rend_real_n = rendimiento_real_por_semana_y_anio(cyc)
    rend_recomendado, cyc_con_resid, trend_stats = rendimiento_recomendado(
        cyc, anio_objetivo, decay=decay, window=window)

    mask_veg_finca = t6m['Referencia'].isin(referencias_vegetal)
    if fincas:
        mask_veg_finca &= t6m['Finca'].isin(fincas)
    n_excluidos = int(t6m[mask_veg_finca]['ciclo_incompleto'].sum())

    return {
        'nombre_vegetal': nombre_vegetal, 'maxcorte': maxcorte, 'anio_objetivo': anio_objetivo,
        'cyc': cyc_con_resid, 'cortes': cortes,
        'curva_tabla': curva_tabla, 'curva_n': curva_n, 'curva_rec': curva_rec,
        'rend_real_tabla': rend_real_tabla, 'rend_real_n': rend_real_n,
        'rend_recomendado': rend_recomendado, 'trend_stats': trend_stats,
        'n_ciclos': len(cyc), 'n_filas_excluidas_incompletas': n_excluidos,
        'anios': sorted(cyc['AñoInicio'].unique().tolist()),
    }


# ---------------------------------------------------------------------------
# Generación del reporte Excel (mismo layout que la pestaña "necesidades",
# más hojas nuevas de curva y rendimiento por año)
# ---------------------------------------------------------------------------
def _styled(cell, value, header=False, sub=False, bold=False, numfmt=None):
    cell.value = value
    cell.font = Font(bold=bold or header, color='FFFFFF' if header else '000000', name=FONT)
    if header:
        cell.fill = H_FILL
    elif sub:
        cell.fill = SUB_FILL
    cell.border = BORDER
    if numfmt:
        cell.number_format = numfmt


def generar_reporte_excel(resultado):
    curva_rec = resultado['curva_rec']
    rend_rec = resultado['rend_recomendado']
    maxcorte = resultado['maxcorte']
    nombre = resultado['nombre_vegetal']
    curva_tabla = resultado['curva_tabla']
    rend_real_tabla = resultado['rend_real_tabla']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'necesidades'

    _styled(ws['B2'], 'Semana de cosecha plan', header=True)
    _styled(ws['C2'], f'Curva de producción {nombre} plan', header=True)
    for i in range(maxcorte):
        wk = i + 1
        _styled(ws[f'B{3+i}'], wk)
        _styled(ws[f'C{3+i}'], round(float(curva_rec[wk]), 3), numfmt='0.0%')
    total_row = 3 + maxcorte
    _styled(ws[f'B{total_row}'], 'Total', bold=True)
    ws[f'C{total_row}'] = f'=SUM(C3:C{total_row-1})'
    ws[f'C{total_row}'].number_format = '0.0%'
    ws[f'C{total_row}'].font = Font(bold=True, name=FONT)

    _styled(ws['E2'], 'Semana', header=True)
    _styled(ws['F2'], 'Rendimiento por semana (recomendado)', header=True)
    for wk in range(1, 54):
        r = 2 + wk
        _styled(ws[f'E{r}'], wk)
        _styled(ws[f'F{r}'], round(float(rend_rec[wk])), numfmt='#,##0')

    # Curva por año
    ws2 = wb.create_sheet('Curva_por_Año')
    _styled(ws2.cell(row=1, column=1), 'Corte', header=True)
    for j, anio in enumerate(curva_tabla.columns):
        _styled(ws2.cell(row=1, column=2+j), int(anio), header=True)
    _styled(ws2.cell(row=1, column=2+len(curva_tabla.columns)), 'Recomendada', header=True)
    for i, corte in enumerate(curva_tabla.index):
        r = i + 2
        _styled(ws2.cell(row=r, column=1), int(corte))
        for j, anio in enumerate(curva_tabla.columns):
            v = curva_tabla.iloc[i, j]
            _styled(ws2.cell(row=r, column=2+j), None if pd.isna(v) else round(float(v), 3), numfmt='0.0%')
        _styled(ws2.cell(row=r, column=2+len(curva_tabla.columns)), round(float(curva_rec[corte]), 3),
                numfmt='0.0%', bold=True)
    for col in range(1, 3+len(curva_tabla.columns)):
        ws2.column_dimensions[get_column_letter(col)].width = 12

    # Rendimiento real por semana y año + recomendado
    ws3 = wb.create_sheet('Rendimiento_por_Año')
    _styled(ws3.cell(row=1, column=1), 'Semana', header=True)
    for j, anio in enumerate(rend_real_tabla.columns):
        _styled(ws3.cell(row=1, column=2+j), int(anio), header=True)
    _styled(ws3.cell(row=1, column=2+len(rend_real_tabla.columns)), 'Recomendado', header=True)
    for i, wk in enumerate(rend_real_tabla.index):
        r = i + 2
        _styled(ws3.cell(row=r, column=1), int(wk))
        for j, anio in enumerate(rend_real_tabla.columns):
            v = rend_real_tabla.iloc[i, j]
            _styled(ws3.cell(row=r, column=2+j), None if pd.isna(v) else round(float(v)), numfmt='#,##0')
        _styled(ws3.cell(row=r, column=2+len(rend_real_tabla.columns)), round(float(rend_rec[wk])),
                numfmt='#,##0', bold=True)
    for col in range(1, 3+len(rend_real_tabla.columns)):
        ws3.column_dimensions[get_column_letter(col)].width = 12

    # Detalle de ciclos (auditoría)
    cyc = resultado['cyc']
    ws4 = wb.create_sheet('Ciclos_Detalle')
    cols = ['Finca', 'Lote', 'Ciclo', 'SemanaInicio', 'AñoInicio', 'Area', 'CantidadV',
            'TotalKilos', 'Rendimiento']
    for i, h in enumerate(cols):
        c = ws4.cell(row=1, column=i+1, value=h)
        c.font = Font(bold=True, color='FFFFFF', name=FONT)
        c.fill = H_FILL
    for i, row in cyc.reset_index(drop=True).iterrows():
        r = i + 2
        for j, col in enumerate(cols):
            v = row[col]
            if col in ('Ciclo', 'SemanaInicio', 'AñoInicio'):
                v = int(v)
            elif col == 'Rendimiento':
                v = round(float(v), 1)
            ws4.cell(row=r, column=j+1, value=v)
    last_row = len(cyc) + 1
    if last_row > 1:
        tab = Table(displayName="TablaCiclos", ref=f"A1:I{last_row}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws4.add_table(tab)
    ws4.freeze_panes = 'A2'

    for col, w in [('B', 14), ('C', 30), ('E', 9), ('F', 26)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
